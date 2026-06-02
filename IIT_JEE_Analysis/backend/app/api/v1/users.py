"""User management endpoints."""
import io
import os
import uuid

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy import select

from app.api.deps import CurrentUser, DbSession, require_roles
from app.crud.user import (
    create_user,
    delete_user,
    get_user_by_email,
    get_user_by_username,
    get_roles,
    get_user,
    get_users,
    update_user,
    update_user_theme,
)
from app.models.mapping import SubjectName
from app.models.user import Role
from app.models.user import RoleName
from app.schemas.student import UploadResult
from app.schemas.user import RoleOut, UserCreate, UserOut, UserThemeUpdate, UserUpdate

router = APIRouter(prefix="/users", tags=["users"])

AVATAR_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "..", "uploads", "avatars")
os.makedirs(AVATAR_DIR, exist_ok=True)

admin_only = Depends(require_roles(RoleName.ADMIN))

_COL_ALIASES: dict[str, str] = {
    "username": "username",
    "user_name": "username",
    "login": "username",
    "full_name": "full_name",
    "name": "full_name",
    "email": "email",
    "phone": "phone",
    "mobile": "phone",
    "whatsapp": "whatsapp",
    "roles": "roles",
    "role": "roles",
    "faculty_subject": "faculty_subject",
    "subject": "faculty_subject",
    "status": "is_active",
    "is_active": "is_active",
    "active": "is_active",
    "password": "password",
}


def _parse_bool(raw: str, default: bool = True) -> bool:
    if not raw:
        return default
    return raw.strip().lower() not in ("inactive", "false", "0", "no", "disabled")


def _parse_roles(raw: str, roles_by_name: dict[str, Role]) -> list[Role]:
    names = [part.strip() for part in raw.replace("|", ",").replace(";", ",").split(",") if part.strip()]
    return [roles_by_name[name.lower()] for name in names]


def _parse_subject(raw: str) -> SubjectName | None:
    if not raw:
        return None
    lowered = raw.strip().lower()
    for subject in SubjectName:
        if subject.value.lower() == lowered:
            return subject
    return None


@router.get("/roles", response_model=list[RoleOut])
def list_roles(db: DbSession, _: CurrentUser):
    return get_roles(db)


@router.get("", response_model=list[UserOut], dependencies=[admin_only])
def list_users(db: DbSession, skip: int = 0, limit: int = 100):
    return get_users(db, skip=skip, limit=limit)


@router.get("/upload/template")
def download_users_template(_: CurrentUser):
    """Return an Excel template for bulk user import."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(500, "openpyxl is not installed on the server")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    headers = ["username", "full_name", "email", "phone", "whatsapp", "roles", "faculty_subject", "status", "password"]
    required = {"username", "full_name", "roles"}
    req_fill = PatternFill("solid", fgColor="1E3A5F")
    opt_fill = PatternFill("solid", fgColor="2E5B8A")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = req_fill if header in required else opt_fill
        cell.alignment = Alignment(horizontal="center")

    ws.append(["faculty01", "Faculty One", "faculty01@example.com", "+91 9876543210", "", "Faculty", "Physics", "Active", "ChangeMe123"])
    ws.append(["operator01", "Operator One", "operator01@example.com", "", "", "Operator", "", "Active", "ChangeMe123"])

    widths = [18, 28, 30, 18, 18, 28, 20, 12, 18]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws2 = wb.create_sheet("Instructions")
    ws2.append(["Column", "Required?", "Notes"])
    ws2.append(["username", "YES", "Unique login name. Existing users are updated on match."])
    ws2.append(["full_name", "YES", "Display name."])
    ws2.append(["email", "no", "Must be unique if provided."])
    ws2.append(["phone", "no", "Mobile number."])
    ws2.append(["whatsapp", "no", "WhatsApp number."])
    ws2.append(["roles", "YES", "Comma-separated. Valid: Admin, Dean, Principal, Vice-Principal, Faculty, Operator."])
    ws2.append(["faculty_subject", "only Faculty", "Required when roles include Faculty. Valid: Mathematics, Physics, Chemistry."])
    ws2.append(["status", "no", "Active or Inactive. Defaults to Active."])
    ws2.append(["password", "new users", "Required for new users. Existing users are changed only if password is filled."])
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="users_template.xlsx"'},
    )


@router.post("/upload/excel", response_model=UploadResult, dependencies=[admin_only])
def upload_users_excel(db: DbSession, file: UploadFile = File(...)):
    """Bulk-upsert users from Excel."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl is not installed on the server")

    try:
        content = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        raise HTTPException(400, f"Cannot read Excel file: {exc}")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Excel file is empty")

    headers = [str(h).strip().lower().replace(" ", "_") if h is not None else "" for h in rows[0]]
    col_map: dict[str, int] = {}
    for idx, header in enumerate(headers):
        canonical = _COL_ALIASES.get(header)
        if canonical and canonical not in col_map:
            col_map[canonical] = idx

    missing = [c for c in ("username", "full_name", "roles") if c not in col_map]
    if missing:
        raise HTTPException(400, f"Missing required columns: {', '.join(missing)}")

    roles_by_name = {role.name.value.lower() if hasattr(role.name, "value") else str(role.name).lower(): role for role in db.scalars(select(Role)).all()}

    created = updated = skipped = 0
    errors: list[str] = []

    for row_num, row in enumerate(rows[1:], start=2):
        def cell(key: str) -> str:
            if key not in col_map:
                return ""
            val = row[col_map[key]]
            return str(val).strip() if val is not None else ""

        try:
            username = cell("username")
            full_name = cell("full_name")
            role_text = cell("roles")
            if not username:
                skipped += 1
                continue
            if not full_name:
                errors.append(f"Row {row_num} ({username}): full_name is empty")
                continue
            if not role_text:
                errors.append(f"Row {row_num} ({username}): roles is empty")
                continue

            role_names = [part.strip() for part in role_text.replace("|", ",").replace(";", ",").split(",") if part.strip()]
            invalid_roles = [name for name in role_names if name.lower() not in roles_by_name]
            if invalid_roles:
                errors.append(f"Row {row_num} ({username}): invalid role(s): {', '.join(invalid_roles)}")
                continue
            roles = _parse_roles(role_text, roles_by_name)

            is_faculty = any((role.name.value if hasattr(role.name, "value") else role.name) == RoleName.FACULTY.value for role in roles)
            subject = _parse_subject(cell("faculty_subject"))
            if is_faculty and not subject:
                errors.append(f"Row {row_num} ({username}): faculty_subject is required for Faculty users")
                continue
            if cell("faculty_subject") and not subject:
                errors.append(f"Row {row_num} ({username}): invalid faculty_subject '{cell('faculty_subject')}'")
                continue

            email = cell("email") or None
            phone = cell("phone") or None
            whatsapp = cell("whatsapp") or None
            is_active = _parse_bool(cell("is_active"))
            password = cell("password")

            existing = get_user_by_username(db, username)
            if email:
                existing_email = get_user_by_email(db, email)
                if existing_email and (not existing or existing_email.id != existing.id):
                    errors.append(f"Row {row_num} ({username}): email '{email}' is already used")
                    continue

            payload = {
                "email": email,
                "full_name": full_name,
                "phone": phone,
                "whatsapp": whatsapp,
                "is_active": is_active,
                "role_ids": [role.id for role in roles],
                "faculty_subjects": [subject] if subject else [],
            }
            if password:
                payload["password"] = password

            if existing:
                update_user(db, existing, UserUpdate(**payload))
                updated += 1
            else:
                if len(password) < 8:
                    errors.append(f"Row {row_num} ({username}): password is required for new users and must be at least 8 characters")
                    continue
                payload.pop("password", None)
                create_user(db, UserCreate(username=username, password=password, **payload))
                created += 1

        except ValueError as exc:
            errors.append(f"Row {row_num}: {exc}")
        except Exception as exc:
            errors.append(f"Row {row_num}: unexpected error - {exc}")

    return UploadResult(created=created, updated=updated, skipped=skipped, errors=errors)


@router.post("", response_model=UserOut, status_code=status.HTTP_201_CREATED, dependencies=[admin_only])
def create_user_endpoint(data: UserCreate, db: DbSession):
    try:
        return create_user(db, data)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.get("/{user_id}", response_model=UserOut)
def get_user_endpoint(user_id: int, db: DbSession, current_user: CurrentUser):
    # Allow self-lookup or admin
    roles = {ur.role.name for ur in current_user.user_roles}
    if current_user.id != user_id and RoleName.ADMIN not in roles:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")
    user = get_user(db, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    return user


@router.patch("/{user_id}", response_model=UserOut, dependencies=[admin_only])
def update_user_endpoint(user_id: int, data: UserUpdate, db: DbSession):
    user = get_user(db, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    try:
        return update_user(db, user, data)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.delete("/{user_id}", status_code=status.HTTP_204_NO_CONTENT, dependencies=[admin_only])
def delete_user_endpoint(user_id: int, db: DbSession):
    user = get_user(db, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    delete_user(db, user)


@router.patch("/me/theme", response_model=UserOut)
def update_my_theme(data: UserThemeUpdate, db: DbSession, current_user: CurrentUser):
    return update_user_theme(db, current_user, data.theme_prefs)


@router.post("/{user_id}/avatar", response_model=UserOut, dependencies=[admin_only])
async def upload_avatar(user_id: int, db: DbSession, file: UploadFile = File(...)):
    user = get_user(db, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    content_type = file.content_type or ""
    if not content_type.startswith("image/"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File must be an image")

    ext = os.path.splitext(file.filename or "")[1] or ".jpg"
    filename = f"{user_id}_{uuid.uuid4().hex}{ext}"
    path = os.path.join(AVATAR_DIR, filename)

    # Remove old avatar file if it exists
    if user.avatar_url:
        old_filename = user.avatar_url.split("/")[-1]
        old_path = os.path.join(AVATAR_DIR, old_filename)
        if os.path.exists(old_path):
            os.remove(old_path)

    contents = await file.read()
    with open(path, "wb") as f:
        f.write(contents)

    from app.schemas.user import UserUpdate as UU
    return update_user(db, user, UU(avatar_url=f"/uploads/avatars/{filename}"))


@router.get("/avatars/{filename}")
def serve_avatar(filename: str):
    path = os.path.join(AVATAR_DIR, filename)
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Not found")
    return FileResponse(path)
