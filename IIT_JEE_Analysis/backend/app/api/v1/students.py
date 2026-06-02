"""Student endpoints — master data + section assignment + Excel bulk upload."""
import io
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import select

from app.api.deps import CurrentUser, DbSession, require_roles
from app.crud.student import (
    assign_section,
    create_student,
    delete_student,
    get_student,
    get_student_by_admission_no,
    get_students,
    get_students_page,
    has_history,
    reactivate_student,
    remove_section,
    update_student,
)
from app.models.branch import Branch
from app.models.class_ import Class
from app.models.mapping import BranchSection
from app.models.program import Program
from app.models.section import Section
from app.models.student import Student
from app.models.student_section import StudentSection
from app.models.user import RoleName
from app.schemas.student import (
    StudentCreate,
    StudentOut,
    StudentPageOut,
    StudentSectionAssign,
    StudentUpdate,
    UploadResult,
)

router = APIRouter(prefix="/students", tags=["students"])
admin_only = Depends(require_roles(RoleName.ADMIN))

_COL_ALIASES: dict[str, str] = {
    "admission_no": "admission_no",
    "adm_no": "admission_no",
    "admno": "admission_no",
    "admission_number": "admission_no",
    "name": "name",
    "student_name": "name",
    "phone": "phone",
    "phone_number": "phone",
    "mobile": "phone",
    "status": "is_active",
    "is_active": "is_active",
    "active": "is_active",
    "target_rank": "target_rank",
    "target_rank": "target_rank",
    "rank": "target_rank",
    "branch_name": "branch_name",
    "branch": "branch_name",
    "program_name": "program_name",
    "program": "program_name",
    "student_class": "class_name",
    "class": "class_name",
    "class_name": "class_name",
    "section": "section_name",
    "section_name": "section_name",
}


# ── Template download ──────────────────────────────────────────────────────────

@router.get("/upload/template")
def download_template(_: CurrentUser):
    """Return a pre-filled .xlsx template the user can populate and upload."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl is not installed on the server")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = ["admission_no", "name", "phone", "status", "target_rank"]
    required = {"admission_no", "name"}

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    opt_fill    = PatternFill("solid", fgColor="2E5B8A")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill if header in required else opt_fill
        cell.alignment = Alignment(horizontal="center")

    # Sample row
    ws.append(["7050729", "CS Koushik", "+91 9876543210", "Active", "Top 100"])

    # Column widths
    widths = [18, 28, 20, 12, 16]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    ws2.append(["Column", "Required?", "Notes"])
    ws2.append(["admission_no",  "YES", "Unique student ID — existing records are updated on match"])
    ws2.append(["name",          "YES", "Full student name"])
    ws2.append(["phone",         "no",  "Mobile number (optional)"])
    ws2.append(["status",        "no",  "Active or Inactive (default: Active)"])
    ws2.append(["target_rank", "no",  "Target Rank — one of: Top 10, Top 100, Top 1000, Top 10000, Qualifier (leave blank for none)"])
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 65

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="students_template.xlsx"'},
    )


# ── Excel bulk upload (static route — must come before /{student_id}) ─────────

@router.post("/upload/excel", response_model=UploadResult, dependencies=[admin_only])
def upload_students_excel(
    db: DbSession,
    file: UploadFile = File(...),
):
    """Bulk-upsert students from Excel. Required columns: admission_no, name. Optional: phone."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl is not installed on the server")

    try:
        content = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        raise HTTPException(400, f"Cannot read Excel file: {exc}")

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise HTTPException(400, "Excel file is empty")

    headers = [
        str(h).strip().lower().replace(" ", "_") if h is not None else ""
        for h in all_rows[0]
    ]
    col_map: dict[str, int] = {}
    for idx, h in enumerate(headers):
        canonical = _COL_ALIASES.get(h)
        if canonical and canonical not in col_map:
            col_map[canonical] = idx

    if "admission_no" not in col_map or "name" not in col_map:
        raise HTTPException(400, "Excel must have at least 'admission_no' and 'name' columns")

    created = updated = skipped = 0
    errors: list[str] = []

    for row_num, row in enumerate(all_rows[1:], start=2):
        def cell(key: str) -> str:
            val = row[col_map[key]]
            return str(val).strip() if val is not None else ""

        try:
            admission_no = cell("admission_no")
            name = cell("name")

            if not admission_no:
                skipped += 1
                continue
            if not name:
                errors.append(f"Row {row_num} ({admission_no}): name is empty")
                continue

            phone = cell("phone") if "phone" in col_map else None
            phone = phone or None

            is_active = True
            if "is_active" in col_map:
                raw = cell("is_active").lower()
                is_active = raw not in ("inactive", "false", "0", "no")

            _VALID_RANKS = {"Top 10", "Top 100", "Top 1000", "Top 10000", "Qualifier"}
            target_rank: str | None = None
            if "target_rank" in col_map:
                raw_rank = cell("target_rank").strip()
                if raw_rank and raw_rank in _VALID_RANKS:
                    target_rank = raw_rank
                elif raw_rank:
                    errors.append(f"Row {row_num} ({admission_no}): invalid target_rank '{raw_rank}' — must be one of {', '.join(_VALID_RANKS)}")
                    continue

            existing = db.scalar(select(Student).where(Student.admission_no == admission_no))
            if existing:
                existing.name = name
                if phone is not None:
                    existing.phone = phone
                existing.is_active = is_active
                if "target_rank" in col_map:
                    existing.target_rank = target_rank
                updated += 1
            else:
                db.add(Student(admission_no=admission_no, name=name, phone=phone, is_active=is_active, target_rank=target_rank))
                created += 1

        except Exception as exc:
            errors.append(f"Row {row_num}: unexpected error — {exc}")

    db.commit()
    return UploadResult(created=created, updated=updated, skipped=skipped, errors=errors)


# ── Section-assignment upload (per year) ──────────────────────────────────────

@router.get("/upload/section-template")
def download_section_template(_: CurrentUser):
    """Template for uploading students with branch/section assignment for a given year."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(500, "openpyxl is not installed on the server")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = ["admission_no", "name", "phone", "status", "target_rank", "branch_name", "program_name", "class_name", "section"]
    required = {"admission_no", "branch_name", "program_name", "class_name", "section"}

    req_fill = PatternFill("solid", fgColor="1E3A5F")
    opt_fill = PatternFill("solid", fgColor="2E5B8A")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = req_fill if header in required else opt_fill
        cell.alignment = Alignment(horizontal="center")

    ws.append(["7050729", "CS Koushik", "+91 9876543210", "Active", "Top 100", "Hyderabad", "IIT-JEE", "Class 11", "A"])

    widths = [18, 28, 20, 12, 16, 20, 18, 14, 10]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws2 = wb.create_sheet("Instructions")
    ws2.append(["Column", "Required?", "Notes"])
    ws2.append(["admission_no",  "YES", "Unique student ID — student is created if not found"])
    ws2.append(["name",          "no",  "Required only when creating a new student"])
    ws2.append(["phone",         "no",  "Mobile number (optional)"])
    ws2.append(["status",        "no",  "Active or Inactive (default: Active)"])
    ws2.append(["target_rank", "no",  "Target Rank — one of: Top 10, Top 100, Top 1000, Top 10000, Qualifier (leave blank for none)"])
    ws2.append(["branch_name",   "YES", "Must exactly match a Branch name in the system"])
    ws2.append(["program_name",  "YES", "Must exactly match a Program name in the system"])
    ws2.append(["class_name",    "YES", "Must exactly match a Class name in the system"])
    ws2.append(["section",       "YES", "Must exactly match a Section name in the system"])
    ws2.append([])
    ws2.append(["Note: the combination of branch / program / class / section must be a configured slot for the selected academic year."])
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="students_section_template.xlsx"'},
    )


@router.post("/upload/section-excel", response_model=UploadResult, dependencies=[admin_only])
def upload_section_excel(
    db: DbSession,
    file: UploadFile = File(...),
    academic_year_id: int = Query(...),
):
    """Upload students with branch/section assignment for a specific academic year.

    Required columns: admission_no, branch_name, program_name, class_name, section.
    name is required only for new students.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl is not installed on the server")

    try:
        content = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        raise HTTPException(400, f"Cannot read Excel file: {exc}")

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise HTTPException(400, "Excel file is empty")

    raw_headers = [
        str(h).strip().lower().replace(" ", "_") if h is not None else ""
        for h in all_rows[0]
    ]
    col_map: dict[str, int] = {}
    for idx, h in enumerate(raw_headers):
        canonical = _COL_ALIASES.get(h)
        if canonical and canonical not in col_map:
            col_map[canonical] = idx

    missing = [c for c in ("admission_no", "branch_name", "program_name", "class_name", "section_name") if c not in col_map]
    if missing:
        raise HTTPException(400, f"Missing required columns: {', '.join(missing)}")

    # Pre-load lookup dicts
    branches = {b.name.lower(): b for b in db.scalars(select(Branch)).all()}
    programs = {p.name.lower(): p for p in db.scalars(select(Program)).all()}
    classes  = {c.name.lower(): c for c in db.scalars(select(Class)).all()}
    sections = {s.name.lower(): s for s in db.scalars(select(Section)).all()}

    # Index branch sections for this year keyed by (branch_id, program_id, class_id, section_id)
    year_bsections = {
        (bs.branch_id, bs.program_id, bs.class_id, bs.section_id): bs
        for bs in db.scalars(
            select(BranchSection).where(BranchSection.academic_year_id == academic_year_id)
        ).all()
    }

    created = updated = skipped = 0
    errors: list[str] = []

    for row_num, row in enumerate(all_rows[1:], start=2):
        def cell(key: str) -> str:
            val = row[col_map[key]]
            return str(val).strip() if val is not None else ""

        try:
            admission_no = cell("admission_no")
            if not admission_no:
                skipped += 1
                continue

            # ── Validate branch / program / class / section ───────────────────
            branch_raw  = cell("branch_name")
            program_raw = cell("program_name")
            class_raw   = cell("class_name")
            section_raw = cell("section_name")

            row_errors: list[str] = []
            branch  = branches.get(branch_raw.lower())
            program = programs.get(program_raw.lower())
            cls     = classes.get(class_raw.lower())
            sec     = sections.get(section_raw.lower())

            if not branch:  row_errors.append(f"branch '{branch_raw}' not found")
            if not program: row_errors.append(f"program '{program_raw}' not found")
            if not cls:     row_errors.append(f"class '{class_raw}' not found")
            if not sec:     row_errors.append(f"section '{section_raw}' not found")

            if row_errors:
                errors.append(f"Row {row_num} ({admission_no}): {'; '.join(row_errors)}")
                continue

            bs_key = (branch.id, program.id, cls.id, sec.id)  # type: ignore[union-attr]
            bs = year_bsections.get(bs_key)
            if not bs:
                errors.append(
                    f"Row {row_num} ({admission_no}): no slot configured for "
                    f"{branch_raw} / {program_raw} / {class_raw} / {section_raw} in this year"
                )
                continue

            # ── Create / update student ───────────────────────────────────────
            phone = cell("phone") if "phone" in col_map else None
            phone = phone or None

            is_active = True
            if "is_active" in col_map:
                raw_status = cell("is_active").lower()
                is_active = raw_status not in ("inactive", "false", "0", "no")

            _VALID_RANKS = {"Top 10", "Top 100", "Top 1000", "Top 10000", "Qualifier"}
            target_rank: str | None = None
            if "target_rank" in col_map:
                raw_rank = cell("target_rank").strip()
                if raw_rank and raw_rank in _VALID_RANKS:
                    target_rank = raw_rank
                elif raw_rank:
                    errors.append(f"Row {row_num} ({admission_no}): invalid target_rank '{raw_rank}' — must be one of {', '.join(sorted(_VALID_RANKS))}")
                    continue

            existing = db.scalar(select(Student).where(Student.admission_no == admission_no))
            if existing:
                if phone is not None:
                    existing.phone = phone
                existing.is_active = is_active
                if "target_rank" in col_map:
                    existing.target_rank = target_rank
                student_id = existing.id
                updated += 1
            else:
                name = cell("name") if "name" in col_map else ""
                if not name:
                    errors.append(f"Row {row_num} ({admission_no}): student not found and 'name' column is empty — cannot create")
                    continue
                s = Student(admission_no=admission_no, name=name, phone=phone, is_active=is_active, target_rank=target_rank)
                db.add(s)
                db.flush()
                student_id = s.id
                created += 1

            # ── Assign section for year ───────────────────────────────────────
            sm = db.scalar(
                select(StudentSection).where(
                    StudentSection.student_id == student_id,
                    StudentSection.academic_year_id == academic_year_id,
                )
            )
            if sm:
                sm.branch_section_id = bs.id
            else:
                db.add(StudentSection(
                    student_id=student_id,
                    branch_section_id=bs.id,
                    academic_year_id=academic_year_id,
                ))

        except Exception as exc:
            errors.append(f"Row {row_num}: unexpected error — {exc}")

    db.commit()
    return UploadResult(created=created, updated=updated, skipped=skipped, errors=errors)


# ── CRUD ──────────────────────────────────────────────────────────────────────

@router.get("", response_model=list[StudentOut])
def list_students(
    db: DbSession,
    _: CurrentUser,
    academic_year_id: Optional[int] = Query(None),
    branch_section_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    skip: int = Query(0),
    limit: int = Query(1000),
):
    return get_students(db, academic_year_id, branch_section_id, search, skip, limit)


@router.get("/page", response_model=StudentPageOut)
def list_students_page(
    db: DbSession,
    _: CurrentUser,
    academic_year_id: Optional[int] = Query(None),
    branch_id: Optional[int] = Query(None),
    branch_ids: Optional[list[int]] = Query(None),
    program_id: Optional[int] = Query(None),
    class_id: Optional[int] = Query(None),
    branch_section_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    sort_field: str = Query("admission_no"),
    sort_dir: str = Query("asc"),
):
    return get_students_page(
        db,
        academic_year_id=academic_year_id,
        branch_id=branch_id,
        branch_ids=branch_ids,
        program_id=program_id,
        class_id=class_id,
        branch_section_id=branch_section_id,
        search=search,
        skip=skip,
        limit=limit,
        sort_field=sort_field,
        sort_dir=sort_dir,
    )


@router.post("", response_model=StudentOut, status_code=201, dependencies=[admin_only])
def create_student_ep(data: StudentCreate, db: DbSession):
    if get_student_by_admission_no(db, data.admission_no):
        raise HTTPException(409, f"Admission no '{data.admission_no}' already exists")
    return create_student(db, data)


@router.get("/{student_id}", response_model=StudentOut)
def get_student_ep(student_id: int, db: DbSession, _: CurrentUser, academic_year_id: Optional[int] = Query(None)):
    obj = get_student(db, student_id, academic_year_id)
    if not obj:
        raise HTTPException(404, "Not found")
    return obj


@router.patch("/{student_id}", response_model=StudentOut, dependencies=[admin_only])
def update_student_ep(student_id: int, data: StudentUpdate, db: DbSession):
    obj = get_student(db, student_id)
    if not obj:
        raise HTTPException(404, "Not found")
    return update_student(db, obj, data)


@router.get("/{student_id}/has-history", dependencies=[admin_only])
def student_has_history(student_id: int, db: DbSession):
    obj = get_student(db, student_id)
    if not obj:
        raise HTTPException(404, "Not found")
    return {"has_history": has_history(db, obj)}


@router.post("/{student_id}/reactivate", response_model=StudentOut, dependencies=[admin_only])
def reactivate_student_ep(student_id: int, db: DbSession):
    obj = get_student(db, student_id)
    if not obj:
        raise HTTPException(404, "Not found")
    if obj.is_active:
        raise HTTPException(400, "Student is already active")
    return reactivate_student(db, obj)


@router.delete("/{student_id}", dependencies=[admin_only])
def delete_student_ep(student_id: int, db: DbSession):
    obj = get_student(db, student_id)
    if not obj:
        raise HTTPException(404, "Not found")
    return delete_student(db, obj)


# ── Section assignment ────────────────────────────────────────────────────────

@router.put("/{student_id}/section", response_model=StudentOut, dependencies=[admin_only])
def assign_section_ep(student_id: int, data: StudentSectionAssign, db: DbSession):
    obj = get_student(db, student_id, data.academic_year_id)
    if not obj:
        raise HTTPException(404, "Student not found")
    bs = db.get(BranchSection, data.branch_section_id)
    if not bs:
        raise HTTPException(404, "Section slot not found")
    return assign_section(db, obj, data.branch_section_id, data.academic_year_id)


@router.delete("/{student_id}/section", response_model=StudentOut, dependencies=[admin_only])
def remove_section_ep(student_id: int, db: DbSession, academic_year_id: int = Query(...)):
    obj = get_student(db, student_id, academic_year_id)
    if not obj:
        raise HTTPException(404, "Student not found")
    return remove_section(db, obj, academic_year_id)
