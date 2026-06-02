"""Exam endpoints — CRUD + per-paper question configuration + OMR results upload."""
import io
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError

from app.api.deps import DbSession, require_roles
from app.crud.exam_result import get_enrolled_students_for_exam, upsert_exam_result
from app.models.exam import Exam, EXAM_STATUS_DRAFT, EXAM_STATUS_PUBLISHED, EXAM_STATUS_COMPLETED
from app.models.exam_question import ExamQuestion
from app.models.mapping import BranchSection
from app.models.student import Student
from app.models.student_section import StudentSection
from app.models.user import RoleName
from app.schemas.exam import ExamCreate, ExamOut, ExamUpdate, ExamQuestionOut, ExamQuestionUpdate, ExamQuestionUploadResult, ExamDetailOut
from app.schemas.exam_result import OMRAbsentStudent, OMRValidationRecord, OMRValidationSummary, OMRUploadConfirm, ExamResultsDetail, StudentResult, QuestionResult

router = APIRouter(prefix="/exams", tags=["exams"])
admin_only = Depends(require_roles(RoleName.ADMIN))

_QUESTION_COLS = [
    "qno", "subject", "topic", "sub_topic", "difficulty",
    "question_type", "marks", "negative_marks", "bkc",
    "partial_marks", "is_deleted", "is_bonus", "akc",
]

_COL_ALIASES: dict[str, str] = {
    "qno": "qno", "q_no": "qno", "question_no": "qno", "question_number": "qno",
    "subject": "subject",
    "topic": "topic",
    "sub_topic": "sub_topic", "subtopic": "sub_topic",
    "difficulty": "difficulty",
    "question_type": "question_type", "type": "question_type", "qtype": "question_type",
    "marks": "marks",
    "negative_marks": "negative_marks", "neg_marks": "negative_marks",
    "bkc": "bkc", "correct_answer": "bkc", "answer": "bkc",
    "partial_marks": "partial_marks",
    "is_deleted": "is_deleted", "deleted": "is_deleted",
    "is_bonus": "is_bonus", "bonus": "is_bonus",
    "akc": "akc", "after_key_change": "akc",
}

_LETTER_TO_NUM = {"A": "1", "B": "2", "C": "3", "D": "4"}
# Question types that take purely numeric answers — letters have no meaning
_NUMERIC_QTYPES = {"INT", "DECIMAL"}
# Question types where multiple options can be selected — expand "BCD" → "2|3|4"
_MULTI_QTYPES = {"MCQ"}
# All recognised question type codes (stored uppercased)
_VALID_QTYPES = {"SCQ", "MCQ", "INT", "DECIMAL"}


def _normalise_answer_key(key: str, question_type: str | None = None) -> str:
    """Convert A/B/C/D option letters to 1/2/3/4 in an answer key, respecting question type.

    SCQ:     single letter or pipe-separated alternates  "A" → "1",  "A|B" → "1|2"
    MCQ:     concatenated letters (no separator in OMR) expanded internally to pipe-separated  "BCD" → "2|3|4"
    INT/DECIMAL: numeric answers, pass through unchanged
    """
    import re
    qt = (question_type or "").strip().upper()

    if qt in _NUMERIC_QTYPES:
        return key

    parts = [p.strip() for p in key.split("|")]
    result: list[str] = []
    for part in parts:
        upper = part.upper()
        if re.fullmatch(r"[A-D]+", upper):
            if qt in _MULTI_QTYPES or len(upper) > 1:
                # Concatenated letters: "BCD" → ["2","3","4"]
                result.extend(_LETTER_TO_NUM[ch] for ch in upper)
            else:
                result.append(_LETTER_TO_NUM.get(upper, part))
        elif qt in _MULTI_QTYPES and re.fullmatch(r"[1-4]{2,}", part):
            # Concatenated digits for MCQ: "234" → ["2","3","4"]
            result.extend(list(part))
        else:
            result.append(part)
    return "|".join(result)


def _siblings(db, exam: Exam) -> list[Exam]:
    """All paper records for the same logical exam (same year + code + program + class)."""
    return list(db.scalars(
        select(Exam).where(
            Exam.academic_year_id == exam.academic_year_id,
            Exam.exam_code == exam.exam_code,
            Exam.program_id == exam.program_id,
            Exam.class_id == exam.class_id,
        ).order_by(Exam.paper)
    ).all())


# ── Exam CRUD ─────────────────────────────────────────────────────────────────

@router.get("", response_model=list[ExamOut])
def list_exams(
    db: DbSession,
    academic_year_id: Optional[int] = Query(None),
    program_id: Optional[int] = Query(None),
    class_id: Optional[int] = Query(None),
):
    stmt = select(Exam)
    if academic_year_id:
        stmt = stmt.where(Exam.academic_year_id == academic_year_id)
    if program_id:
        stmt = stmt.where(Exam.program_id == program_id)
    if class_id:
        stmt = stmt.where(Exam.class_id == class_id)
    stmt = stmt.order_by(Exam.exam_date, Exam.exam_code, Exam.paper)
    exams = db.scalars(stmt).all()

    counts: dict[int, int] = {}
    result_counts: dict[int, int] = {}
    if exams:
        exam_ids = [e.id for e in exams]
        count_rows = db.execute(
            select(ExamQuestion.exam_id, func.count(ExamQuestion.id).label("cnt"))
            .where(
                ExamQuestion.exam_id.in_(exam_ids),
                ExamQuestion.is_deleted == False,  # noqa: E712
            )
            .group_by(ExamQuestion.exam_id)
        ).all()
        counts = {r.exam_id: r.cnt for r in count_rows}

        from app.models.exam_result import ExamResult
        from app.models.exam_upload_log import ExamUploadLog
        from app.schemas.exam_result import UploadLogOut

        result_rows = db.execute(
            select(ExamResult.exam_id, func.count(ExamResult.id).label("cnt"))
            .where(ExamResult.exam_id.in_(exam_ids))
            .group_by(ExamResult.exam_id)
        ).all()
        result_counts = {r.exam_id: r.cnt for r in result_rows}

        log_rows = db.scalars(
            select(ExamUploadLog).where(ExamUploadLog.exam_id.in_(exam_ids))
        ).all()
        # Group logs by exam_id → list
        logs_by_exam: dict[int, list] = {}
        for log in log_rows:
            logs_by_exam.setdefault(log.exam_id, []).append(log)
    else:
        from app.schemas.exam_result import UploadLogOut
        logs_by_exam = {}

    def _log_out(log):
        return UploadLogOut(
            branch_id=log.branch_id,
            uploaded_at=log.uploaded_at,
            valid_count=log.valid_count,
            absent_count=log.absent_count,
            duplicate_count=log.duplicate_count,
            invalid_count=log.invalid_count,
            absent_list=[a for a in log.absent_list.split(",") if a],
            file_name=log.file_name,
        )

    return [
        ExamOut.model_validate(e).model_copy(update={
            "question_count": counts.get(e.id, 0),
            "result_count": result_counts.get(e.id, 0),
            "upload_logs": [_log_out(lg) for lg in logs_by_exam.get(e.id, [])],
        })
        for e in exams
    ]


@router.get("/{exam_id}/detail", response_model=ExamDetailOut)
def get_exam_detail(exam_id: int, db: DbSession):
    """Get detailed exam view with ALL branches; sections/students shown where configured."""
    exam = db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(404, "Exam not found")

    from sqlalchemy.orm import selectinload
    from app.models.branch import Branch
    from app.models.mapping import DeanBranch, PrincipalBranch
    from app.models.user import User

    # All active branches
    all_branches = db.scalars(
        select(Branch).where(Branch.is_active == True).order_by(Branch.name)  # noqa: E712
    ).all()

    # Branch sections configured for this exam's program/class/year
    branch_sections = db.scalars(
        select(BranchSection).where(
            BranchSection.academic_year_id == exam.academic_year_id,
            BranchSection.program_id == exam.program_id,
            BranchSection.class_id == exam.class_id,
        ).options(
            selectinload(BranchSection.section),
        )
    ).all()

    # Group branch sections by branch_id
    bs_by_branch: dict[int, list] = {}
    for bs in branch_sections:
        bs_by_branch.setdefault(bs.branch_id, []).append(bs)

    # Students for all branch sections
    students_by_bs: dict[int, list] = {}
    if branch_sections:
        student_sections = db.scalars(
            select(StudentSection).where(
                StudentSection.academic_year_id == exam.academic_year_id,
                StudentSection.branch_section_id.in_([bs.id for bs in branch_sections]),
            ).options(
                selectinload(StudentSection.student),
            )
        ).all()
        for ss in student_sections:
            students_by_bs.setdefault(ss.branch_section_id, []).append(ss.student)

    # Leadership per branch (batch fetch)
    dean_rows = db.execute(
        select(DeanBranch.branch_id, User.id, User.full_name, User.username)
        .join(User, DeanBranch.user_id == User.id)
    ).all()
    dean_by_branch = {
        r.branch_id: {"id": r.id, "full_name": r.full_name, "username": r.username}
        for r in dean_rows
    }

    principal_rows = db.execute(
        select(PrincipalBranch.branch_id, User.id, User.full_name, User.username)
        .join(User, PrincipalBranch.user_id == User.id)
    ).all()
    principal_by_branch = {
        r.branch_id: {"id": r.id, "full_name": r.full_name, "username": r.username}
        for r in principal_rows
    }

    branches_detail = []
    for branch in all_branches:
        sections = []
        for bs in bs_by_branch.get(branch.id, []):
            students = students_by_bs.get(bs.id, [])
            sections.append({
                "section_name": bs.section.name,
                "student_count": len(students),
                "students": [
                    {"id": s.id, "admission_no": s.admission_no, "name": s.name}
                    for s in students
                ],
            })
        sections.sort(key=lambda s: s["section_name"])

        branches_detail.append({
            "id": branch.id,
            "name": branch.name,
            "code": branch.code,
            "address": branch.address,
            "principal": principal_by_branch.get(branch.id),
            "dean": dean_by_branch.get(branch.id),
            "operator": None,
            "sections": sections,
        })

    return {
        "id": exam.id,
        "exam_code": exam.exam_code,
        "exam_type": exam.exam_type,
        "exam_date": exam.exam_date,
        "program_name": exam.program.name,
        "class_name": exam.class_.name,
        "branches": branches_detail,
    }


@router.post("", response_model=list[ExamOut], dependencies=[admin_only])
def create_exam(data: ExamCreate, db: DbSession):
    papers = ["P1"] if data.exam_type == "Mains" else ["P1", "P2"]
    records = [
        Exam(
            academic_year_id=data.academic_year_id,
            program_id=data.program_id,
            class_id=data.class_id,
            exam_code=data.exam_code,
            exam_type=data.exam_type,
            paper=p,
            exam_date=data.exam_date,
            mas_mathematics=data.mas_mathematics,
            mas_physics=data.mas_physics,
            mas_chemistry=data.mas_chemistry,
        )
        for p in papers
    ]
    for r in records:
        db.add(r)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(409, f"Exam '{data.exam_code}' already exists for this program/class/year")
    for r in records:
        db.refresh(r)
    return records


_MAS_FIELDS = {"mas_mathematics", "mas_physics", "mas_chemistry"}

@router.patch("/{exam_id}", response_model=list[ExamOut], dependencies=[admin_only])
def update_exam(exam_id: int, data: ExamUpdate, db: DbSession):
    exam = db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(404, "Exam not found")
    updates = data.model_dump(exclude_unset=True)
    non_mas = {k for k in updates if k not in _MAS_FIELDS}
    if non_mas and exam.status != EXAM_STATUS_DRAFT:
        raise HTTPException(409, "Only draft exams can be edited")
    if exam.status == EXAM_STATUS_COMPLETED:
        raise HTTPException(409, "Completed exams cannot be modified")
    # MAS patch: update only the specific paper, not siblings
    if updates.keys() <= _MAS_FIELDS:
        for k, v in updates.items():
            setattr(exam, k, v)
    else:
        for sib in _siblings(db, exam):
            for k, v in updates.items():
                setattr(sib, k, v)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(409, "Exam already exists for this program/class/year")
    return _siblings(db, exam)


@router.delete("/{exam_id}", dependencies=[admin_only])
def delete_exam(exam_id: int, db: DbSession):
    exam = db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(404, "Exam not found")
    if exam.status == EXAM_STATUS_PUBLISHED:
        raise HTTPException(409, "Published exams cannot be deleted — unpublish first")

    sibling_ids = [s.id for s in _siblings(db, exam)]

    # Use a raw table DELETE so DB-level CASCADE handles all child rows
    # (exam_results, exam_questions, exam_upload_log, student_evaluations)
    db.execute(Exam.__table__.delete().where(Exam.id.in_(sibling_ids)))
    db.commit()
    return {"ok": True}


# ── Lifecycle endpoints ────────────────────────────────────────────────────────

@router.post("/{exam_id}/publish", response_model=ExamOut, dependencies=[admin_only])
def publish_exam(exam_id: int, db: DbSession):
    exam = db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(404, "Exam not found")
    if exam.status != EXAM_STATUS_DRAFT:
        raise HTTPException(409, "Only draft exams can be published")
    q_count = db.scalar(
        select(func.count(ExamQuestion.id)).where(
            ExamQuestion.exam_id == exam_id,
            ExamQuestion.is_deleted == False,  # noqa: E712
        )
    )
    if not q_count:
        raise HTTPException(409, "Cannot publish — no questions configured for this paper")
    for sib in _siblings(db, exam):
        sib.status = EXAM_STATUS_PUBLISHED
    db.commit()
    db.refresh(exam)
    return ExamOut.model_validate(exam).model_copy(update={"question_count": q_count})


@router.post("/{exam_id}/unpublish", response_model=ExamOut, dependencies=[admin_only])
def unpublish_exam(exam_id: int, db: DbSession):
    from app.models.exam_result import ExamResult
    exam = db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(404, "Exam not found")
    if exam.status != EXAM_STATUS_PUBLISHED:
        raise HTTPException(409, "Only published exams can be unpublished")
    has_results = db.scalar(
        select(func.count(ExamResult.id)).where(ExamResult.exam_id == exam_id)
    )
    if has_results:
        raise HTTPException(409, "Cannot unpublish — OMR results have already been uploaded")
    for sib in _siblings(db, exam):
        sib.status = EXAM_STATUS_DRAFT
    db.commit()
    db.refresh(exam)
    q_count = db.scalar(
        select(func.count(ExamQuestion.id)).where(
            ExamQuestion.exam_id == exam_id, ExamQuestion.is_deleted == False  # noqa: E712
        )
    )
    return ExamOut.model_validate(exam).model_copy(update={"question_count": q_count or 0})


@router.post("/{exam_id}/complete", response_model=ExamOut, dependencies=[admin_only])
def complete_exam(exam_id: int, db: DbSession):
    exam = db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(404, "Exam not found")
    if exam.status != EXAM_STATUS_PUBLISHED:
        raise HTTPException(409, "Only published exams can be marked as completed")
    for sib in _siblings(db, exam):
        sib.status = EXAM_STATUS_COMPLETED
    db.commit()
    db.refresh(exam)
    q_count = db.scalar(
        select(func.count(ExamQuestion.id)).where(
            ExamQuestion.exam_id == exam_id, ExamQuestion.is_deleted == False  # noqa: E712
        )
    )
    return ExamOut.model_validate(exam).model_copy(update={"question_count": q_count or 0})


@router.post("/{exam_id}/reopen", response_model=ExamOut, dependencies=[admin_only])
def reopen_exam(exam_id: int, db: DbSession):
    """Move a completed exam back to published so results can still be uploaded."""
    exam = db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(404, "Exam not found")
    if exam.status != EXAM_STATUS_COMPLETED:
        raise HTTPException(409, "Only completed exams can be reopened")
    for sib in _siblings(db, exam):
        sib.status = EXAM_STATUS_PUBLISHED
    db.commit()
    db.refresh(exam)
    q_count = db.scalar(
        select(func.count(ExamQuestion.id)).where(
            ExamQuestion.exam_id == exam_id, ExamQuestion.is_deleted == False  # noqa: E712
        )
    )
    return ExamOut.model_validate(exam).model_copy(update={"question_count": q_count or 0})


@router.delete("/{exam_id}/results", dependencies=[admin_only])
def clear_exam_results(
    exam_id: int,
    db: DbSession,
    branch_id: Optional[int] = Query(None, description="If provided, clear only this branch's results"),
):
    """Delete OMR results and upload log. If branch_id given, clears only that branch."""
    from app.models.exam_result import ExamResult
    from app.models.exam_upload_log import ExamUploadLog
    from app.models.student_section import StudentSection
    from app.models.mapping import BranchSection
    exam = db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(404, "Exam not found")

    if branch_id is not None:
        # Delete results only for students enrolled in this branch
        student_ids = db.scalars(
            select(StudentSection.student_id)
            .join(BranchSection, StudentSection.branch_section_id == BranchSection.id)
            .where(
                StudentSection.academic_year_id == exam.academic_year_id,
                BranchSection.program_id == exam.program_id,
                BranchSection.class_id == exam.class_id,
                BranchSection.branch_id == branch_id,
            )
        ).all()
        deleted = db.execute(
            ExamResult.__table__.delete().where(
                ExamResult.exam_id == exam_id,
                ExamResult.student_id.in_(student_ids),
            )
        )
        db.execute(
            ExamUploadLog.__table__.delete().where(
                ExamUploadLog.exam_id == exam_id,
                ExamUploadLog.branch_id == branch_id,
            )
        )
    else:
        deleted = db.execute(
            ExamResult.__table__.delete().where(ExamResult.exam_id == exam_id)
        )
        db.execute(
            ExamUploadLog.__table__.delete().where(ExamUploadLog.exam_id == exam_id)
        )
    db.commit()
    return {"deleted": deleted.rowcount}


@router.post("/{exam_id}/evaluate", dependencies=[admin_only])
def evaluate_exam(exam_id: int, db: DbSession):
    """
    Run evaluation for a completed exam — computes ranks, MI, percentile, and
    persists StudentEvaluation (per paper) and StudentCumulativeEvaluation (Advanced only).
    Can be re-run to refresh results.
    """
    from app.crud.evaluation import run_evaluation
    from app.models.student_evaluation import StudentEvaluation

    exam = db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(404, "Exam not found")
    if exam.status != EXAM_STATUS_COMPLETED:
        raise HTTPException(409, "Evaluation can only be run on completed exams")

    # Check at least one paper has results
    from app.models.exam_result import ExamResult
    has_results = db.scalar(
        select(func.count(ExamResult.id)).where(
            ExamResult.exam_id.in_(
                select(Exam.id).where(
                    Exam.academic_year_id == exam.academic_year_id,
                    Exam.exam_code == exam.exam_code,
                    Exam.program_id == exam.program_id,
                    Exam.class_id == exam.class_id,
                )
            )
        )
    )
    if not has_results:
        raise HTTPException(409, "No OMR results uploaded — upload results before evaluating")

    summary = run_evaluation(db, exam)
    return summary


@router.get("/{exam_id}/evaluation/status")
def get_evaluation_status(exam_id: int, db: DbSession):
    """Return whether this exam has been evaluated and when."""
    from app.models.student_evaluation import StudentEvaluation, StudentCumulativeEvaluation
    from sqlalchemy import func as sqlfunc

    exam = db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(404, "Exam not found")

    eval_row = db.execute(
        select(
            sqlfunc.count(StudentEvaluation.id).label("count"),
            sqlfunc.max(StudentEvaluation.evaluated_at).label("last_evaluated_at"),
        ).where(StudentEvaluation.exam_id == exam_id)
    ).one()

    cumul_count = db.scalar(
        select(sqlfunc.count(StudentCumulativeEvaluation.id))
        .where(StudentCumulativeEvaluation.p1_exam_id == exam_id)
    ) or 0

    return {
        "evaluated": eval_row.count > 0,
        "evaluated_count": eval_row.count,
        "last_evaluated_at": eval_row.last_evaluated_at,
        "cumulative_count": cumul_count,
    }


# ── Per-paper question endpoints ──────────────────────────────────────────────

@router.delete("/{exam_id}/questions", dependencies=[admin_only])
def clear_exam_questions(exam_id: int, db: DbSession):
    """Delete ALL questions for a paper. Only allowed when exam is in draft status."""
    exam = db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(404, "Exam not found")
    if exam.status != EXAM_STATUS_DRAFT:
        raise HTTPException(409, "Questions can only be cleared from draft exams")
    deleted = db.execute(
        ExamQuestion.__table__.delete().where(ExamQuestion.exam_id == exam_id)
    )
    db.commit()
    return {"deleted": deleted.rowcount}


@router.get("/{exam_id}/questions", response_model=list[ExamQuestionOut])
def list_exam_questions(exam_id: int, db: DbSession, include_deleted: bool = Query(False)):
    if not db.get(Exam, exam_id):
        raise HTTPException(404, "Exam not found")
    stmt = select(ExamQuestion).where(ExamQuestion.exam_id == exam_id)
    if not include_deleted:
        stmt = stmt.where(ExamQuestion.is_deleted == False)  # noqa: E712
    stmt = stmt.order_by(ExamQuestion.qno)
    return db.scalars(stmt).all()


@router.patch("/{exam_id}/questions/{question_id}", response_model=ExamQuestionOut, dependencies=[admin_only])
def update_exam_question(exam_id: int, question_id: int, data: ExamQuestionUpdate, db: DbSession):
    exam = db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(404, "Exam not found")
    q = db.get(ExamQuestion, question_id)
    if not q or q.exam_id != exam_id:
        raise HTTPException(404, "Question not found")

    updates = data.model_dump(exclude_none=True)

    # Normalise A/B/C/D → 1/2/3/4 in answer keys, using the effective question type
    effective_qtype = updates.get("question_type") or q.question_type
    for key_field in ("bkc", "akc"):
        if key_field in updates and updates[key_field]:
            updates[key_field] = _normalise_answer_key(updates[key_field], effective_qtype)

    # When published or completed only AKC and flags can be changed
    if exam.status in (EXAM_STATUS_PUBLISHED, EXAM_STATUS_COMPLETED):
        allowed = {"akc", "is_deleted", "is_bonus"}
        disallowed = {k for k in updates if k not in allowed}
        if disallowed:
            raise HTTPException(409, "Exam is published — only AKC and flags can be updated")

    # Completed exams are fully locked
    if exam.status == EXAM_STATUS_COMPLETED:
        raise HTTPException(409, "Exam is completed — no further changes allowed")

    for k, v in updates.items():
        setattr(q, k, v)
    db.commit()
    db.refresh(q)
    return q


@router.get("/{exam_id}/questions/upload/template")
def download_exam_question_template(exam_id: int, db: DbSession):
    if not db.get(Exam, exam_id):
        raise HTTPException(404, "Exam not found")
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws.append(_QUESTION_COLS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DCE6F1")
        cell.alignment = Alignment(horizontal="center")
    ws.append([1, "Mathematics", "Algebra", "Quadratic Equations", "Medium", "SCQ", 4, 1, "A|B", 0, False, False, "C"])
    ws.append([2, "Physics", "Mechanics", "Newton's Laws", "Hard", "INT", 4, 1, "1|2", 0, False, False, "3"])
    ws.append([3, "Chemistry", "Organic", "Hydrocarbons", "Easy", "DECIMAL", 4, 1, "3.14:3.15", 0, False, False, ""])
    ws.append([4, "Mathematics", "Calculus", "Integration", "Very Hard", "MCQ", 4, 2, "ABD", 1, False, False, "ACD"])
    ws.append([5, "Physics", "Optics", "Refraction", "Easy", "SCQ", 4, 1, "C", 0, False, True, ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=exam_questions_template.xlsx"},
    )


@router.post("/{exam_id}/questions/upload/excel", response_model=ExamQuestionUploadResult, dependencies=[admin_only])
def upload_exam_questions(exam_id: int, db: DbSession, file: UploadFile = File(...)):
    exam = db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(404, "Exam not found")
    if exam.status != EXAM_STATUS_DRAFT:
        raise HTTPException(409, "Questions can only be uploaded to draft exams")

    import openpyxl

    created = updated = skipped = 0
    errors: list[str] = []

    try:
        wb = openpyxl.load_workbook(file.file, data_only=True)
    except Exception:
        raise HTTPException(400, "Invalid Excel file")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Empty file")

    raw_headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]
    col_map: dict[str, int] = {}
    for i, h in enumerate(raw_headers):
        canon = _COL_ALIASES.get(h)
        if canon and canon not in col_map:
            col_map[canon] = i

    if "qno" not in col_map or "subject" not in col_map:
        raise HTTPException(400, "Missing required columns: qno, subject")

    existing = {
        q.qno: q
        for q in db.scalars(select(ExamQuestion).where(ExamQuestion.exam_id == exam_id)).all()
    }

    def _str(val) -> Optional[str]:
        return str(val).strip() if val is not None and str(val).strip() else None

    def _float(val) -> Optional[float]:
        if val is None or str(val).strip() == "":
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    def _bool(val) -> bool:
        if isinstance(val, bool):
            return val
        return str(val).strip().lower() in ("true", "1", "yes", "deleted")

    # ── Pass 1: validate all rows, collect parsed data ────────────────────
    seen_qnos: set[int] = set()
    validated_rows: list[dict] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            qno_raw = row[col_map["qno"]] if "qno" in col_map else None
            if qno_raw is None or str(qno_raw).strip() == "":
                skipped += 1
                continue
            try:
                qno = int(float(qno_raw))
            except (ValueError, TypeError):
                errors.append(f"Row {row_idx}: qno '{qno_raw}' is not a valid integer")
                skipped += 1
                continue

            if qno in seen_qnos:
                errors.append(f"Row {row_idx}: qno {qno} is duplicated in this file")
                skipped += 1
                continue
            seen_qnos.add(qno)

            subject = _str(row[col_map["subject"]]) if "subject" in col_map else None
            if not subject:
                errors.append(f"Row {row_idx}: subject is required")
                skipped += 1
                continue

            raw_qtype = _str(row[col_map["question_type"]]) if "question_type" in col_map else None
            question_type: str | None = None
            if raw_qtype:
                normalised_qt = raw_qtype.strip().upper()
                if normalised_qt not in _VALID_QTYPES:
                    errors.append(
                        f"Row {row_idx}: invalid question_type '{raw_qtype}'. "
                        f"Must be one of: {', '.join(sorted(_VALID_QTYPES))}"
                    )
                    skipped += 1
                    continue
                question_type = normalised_qt

            validated_rows.append({
                "qno": qno,
                "subject": subject,
                "topic": _str(row[col_map["topic"]]) if "topic" in col_map else None,
                "sub_topic": _str(row[col_map["sub_topic"]]) if "sub_topic" in col_map else None,
                "difficulty": _str(row[col_map["difficulty"]]) if "difficulty" in col_map else None,
                "question_type": question_type,
                "marks": _float(row[col_map["marks"]]) if "marks" in col_map else None,
                "negative_marks": _float(row[col_map["negative_marks"]]) if "negative_marks" in col_map else None,
                "bkc": (_normalise_answer_key(v, question_type) if (v := _str(row[col_map["bkc"]])) else None) if "bkc" in col_map else None,
                "partial_marks": _float(row[col_map["partial_marks"]]) if "partial_marks" in col_map else None,
                "is_deleted": _bool(row[col_map["is_deleted"]]) if "is_deleted" in col_map and row[col_map["is_deleted"]] is not None else False,
                "is_bonus": _bool(row[col_map["is_bonus"]]) if "is_bonus" in col_map and row[col_map["is_bonus"]] is not None else False,
                "akc": (_normalise_answer_key(v, question_type) if (v := _str(row[col_map["akc"]])) else None) if "akc" in col_map else None,
            })

        except Exception as exc:
            errors.append(f"Row {row_idx}: {exc}")
            skipped += 1

    # ── If any errors, abort — nothing is saved ────────────────────────────
    if errors:
        return ExamQuestionUploadResult(created=0, updated=0, skipped=skipped, errors=errors)

    # ── Pass 2: all rows valid — apply to DB and commit ───────────────────
    for data in validated_rows:
        qno = data["qno"]
        if qno in existing:
            q = existing[qno]
            for k, v in data.items():
                setattr(q, k, v)
            updated += 1
        else:
            q = ExamQuestion(exam_id=exam_id, **data)
            db.add(q)
            existing[qno] = q
            created += 1

    db.commit()
    return ExamQuestionUploadResult(created=created, updated=updated, skipped=skipped, errors=errors)


# ── OMR results upload ────────────────────────────────────────────────────────

@router.get("/{exam_id}/results", response_model=ExamResultsDetail)
def get_exam_results(exam_id: int, db: DbSession):
    """Return per-student scored responses computed on the fly from current question keys."""
    from app.models.exam_result import ExamResult

    exam = db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(404, "Exam not found")

    # All questions for this paper (including deleted — needed to align answer positions)
    questions = db.scalars(
        select(ExamQuestion).where(ExamQuestion.exam_id == exam_id).order_by(ExamQuestion.qno)
    ).all()

    # All saved results
    results = db.scalars(
        select(ExamResult)
        .where(ExamResult.exam_id == exam_id)
        .order_by(ExamResult.student_id)
    ).all()

    # Fetch student info + branch in one pass
    student_ids = [r.student_id for r in results]
    from app.models.student import Student
    from app.models.student_section import StudentSection
    from app.models.branch import Branch

    students_map: dict[int, Student] = {}
    student_branch: dict[int, tuple[int, str]] = {}  # student_id -> (branch_id, branch_name)
    if student_ids:
        for s in db.scalars(select(Student).where(Student.id.in_(student_ids))).all():
            students_map[s.id] = s

        # Fetch branch via StudentSection → BranchSection → Branch
        rows = db.execute(
            select(StudentSection.student_id, BranchSection.branch_id, Branch.name)
            .join(BranchSection, StudentSection.branch_section_id == BranchSection.id)
            .join(Branch, BranchSection.branch_id == Branch.id)
            .where(
                StudentSection.student_id.in_(student_ids),
                StudentSection.academic_year_id == exam.academic_year_id,
            )
        ).all()
        for row in rows:
            student_branch[row.student_id] = (row.branch_id, row.name)

    # Build question meta list (0-indexed by position in sorted qno order)
    question_meta = [
        {
            "qno": q.qno,
            "subject": q.subject,
            "question_type": q.question_type,
            "marks": q.marks or 0,
            "negative_marks": q.negative_marks or 0,
            "bkc": q.bkc,
            "akc": q.akc,
            "is_bonus": q.is_bonus,
            "is_deleted": q.is_deleted,
        }
        for q in questions
    ]

    def _parse_answers(answers_csv: str) -> list[int | str]:
        """Parse stored answer CSV back into typed values, using question type per position."""
        result: list[int | str] = []
        for i, a in enumerate(answers_csv.split(",")):
            a = a.strip()
            if not a:
                continue
            qt = (questions[i].question_type or "").upper() if i < len(questions) else ""
            if "|" in a:
                result.append(a)
            elif qt == "MCQ":
                # Normalise concatenated letters/digits: "BCD"→"2|3|4", "234"→"2|3|4"
                result.append(_normalise_answer_key(a, "MCQ"))
            elif qt == "DECIMAL":
                result.append(a)
            else:
                try:
                    result.append(int(a))
                except ValueError:
                    result.append(0)
        return result

    def _is_unattempted(ans: int | str, qtype: str | None = None) -> bool:
        # -1000000 is the OMR sentinel for absent/unattempted
        if ans == -1000000:
            return True
        qt = (qtype or "").upper()
        # For INT, 0 is a valid answer (digit 0–9), so only -1000000 means unattempted
        if qt == "INT":
            return False
        # For SCQ/MCQ/DECIMAL: 0 or "0" means no selection (unattempted)
        return ans == 0 or str(ans).strip() == "0"

    def _key_options(key: str) -> frozenset[str]:
        """Split a pipe-separated key into a set of accepted option strings."""
        return frozenset(p.strip() for p in key.split("|") if p.strip())

    def _score_question(ans: int | str, key: str, qtype: str | None, marks: float, neg: float, partial: float) -> tuple[float, bool | None]:
        """
        Returns (marks_awarded, is_correct).
        is_correct: True=full correct, False=wrong, None=partial.

        SCQ     — option 1-4; key is pipe-separated alternates e.g. "1|2|4"; correct → +marks, wrong → -neg
        INT     — integer 0-9; key is pipe-separated alternates e.g. "1|2"; correct → +marks, wrong → -neg
        DECIMAL — decimal value; key is colon-separated range e.g. "3.14:3.15"; correct → +marks, wrong → -neg
        MCQ     — pipe-separated selected options e.g. "1|3"; key is pipe-separated correct set;
                  full match → +marks, any wrong option selected → -neg, correct subset only → +partial×n
        """
        key_str = key.strip()
        qt = (qtype or "").upper()
        ans_str = str(ans).strip()

        if qt == "SCQ":
            accepted = _key_options(key_str)
            if ans_str in accepted:
                return marks, True
            return -neg, False

        if qt == "INT":
            # Key is pipe-separated alternate accepted integer values
            accepted = _key_options(key_str)
            if ans_str in accepted:
                return marks, True
            # Also try numeric equality for keys stored as floats e.g. "1.0"
            try:
                ans_int = str(int(float(ans_str)))
                if ans_int in accepted:
                    return marks, True
                for k in accepted:
                    if int(float(k)) == int(float(ans_str)):
                        return marks, True
            except (ValueError, TypeError):
                pass
            return -neg, False

        if qt == "DECIMAL":
            # Key is colon-separated range e.g. "3.14:3.15"; wrong answer → -neg
            try:
                val = float(ans_str)
            except ValueError:
                return -neg, False
            if ":" in key_str:
                try:
                    lo_s, hi_s = key_str.split(":", 1)
                    lo, hi = float(lo_s.strip()), float(hi_s.strip())
                    if min(lo, hi) <= val <= max(lo, hi):
                        return marks, True
                except ValueError:
                    pass
            else:
                try:
                    if float(key_str) == val:
                        return marks, True
                except ValueError:
                    pass
            return -neg, False

        if qt == "MCQ":
            selected = _key_options(ans_str)
            correct = _key_options(key_str)
            if selected == correct:
                return marks, True
            wrong_selected = selected - correct
            if wrong_selected:
                return -neg, False
            # Correct subset only (no wrong selections) → partial per correct option chosen
            n = len(selected & correct)
            return n * (partial if partial else 1.0), None

        return 0.0, False

    def score_student(answers_csv: str):
        student_answers = _parse_answers(answers_csv)

        responses: list[QuestionResult] = []
        total = math_s = phys_s = chem_s = 0.0
        attempted = correct = wrong = unattempted = 0

        for i, q in enumerate(questions):
            ans: int | str = student_answers[i] if i < len(student_answers) else 0
            effective_key = q.akc if q.akc else q.bkc
            marks = float(q.marks or 0)
            neg = float(q.negative_marks or 0)
            partial = float(q.partial_marks or 0)

            if q.is_deleted:
                awarded = 0.0
                is_correct = None
            elif q.is_bonus:
                awarded = marks
                is_correct = None
            elif _is_unattempted(ans, q.question_type):
                awarded = 0.0
                is_correct = None
                unattempted += 1
            elif not effective_key:
                awarded = 0.0
                is_correct = None
            else:
                awarded, is_correct = _score_question(ans, effective_key, q.question_type, marks, neg, partial)
                if is_correct is True:
                    correct += 1
                    attempted += 1
                elif is_correct is False:
                    wrong += 1
                    attempted += 1
                else:
                    # partial
                    attempted += 1

            if not q.is_deleted:
                total += awarded
                if q.subject == "Mathematics":
                    math_s += awarded
                elif q.subject == "Physics":
                    phys_s += awarded
                elif q.subject == "Chemistry":
                    chem_s += awarded

            responses.append(QuestionResult(
                qno=q.qno,
                subject=q.subject,
                question_type=q.question_type,
                student_answer=ans,
                correct_answer=effective_key,
                is_correct=is_correct,
                marks_awarded=awarded,
                is_bonus=q.is_bonus,
                is_deleted=q.is_deleted,
            ))

        return responses, total, math_s, phys_s, chem_s, attempted, correct, wrong, unattempted

    student_results: list[StudentResult] = []
    for result in results:
        student = students_map.get(result.student_id)
        if not student:
            continue
        responses, total, math_s, phys_s, chem_s, attempted, correct, wrong, unattempted = score_student(result.answers)
        branch_info = student_branch.get(student.id)
        student_results.append(StudentResult(
            student_id=student.id,
            admission_no=student.admission_no,
            name=student.name,
            target_rank=student.target_rank,
            branch_id=branch_info[0] if branch_info else None,
            branch_name=branch_info[1] if branch_info else None,
            total_score=total,
            math_score=math_s,
            physics_score=phys_s,
            chemistry_score=chem_s,
            attempted=attempted,
            correct=correct,
            wrong=wrong,
            unattempted=unattempted,
            responses=responses,
        ))

    # Sort by total score descending
    student_results.sort(key=lambda s: s.total_score, reverse=True)

    return ExamResultsDetail(
        exam_id=exam_id,
        questions=question_meta,
        students=student_results,
    )


@router.post("/{exam_id}/results/validate", response_model=OMRValidationSummary, dependencies=[admin_only])
def validate_omr_file(
    exam_id: int,
    db: DbSession,
    branch_id: int = Query(..., description="Branch to filter enrolled students"),
    file: UploadFile = File(...),
):
    """Validate OMR file and return validation summary without saving."""
    exam = db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(404, "Exam not found")
    if exam.status != EXAM_STATUS_PUBLISHED:
        raise HTTPException(409, "OMR results can only be uploaded to published exams")

    # Load questions to know type per position
    questions = db.scalars(
        select(ExamQuestion).where(ExamQuestion.exam_id == exam_id).order_by(ExamQuestion.qno)
    ).all()
    qtypes = [(q.question_type or "").upper() for q in questions]

    # Read file
    try:
        content = file.file.read().decode("utf-8")
    except Exception:
        raise HTTPException(400, "Cannot read file — ensure it's plain text CSV")

    lines = [line.strip() for line in content.split("\n") if line.strip()]
    if not lines:
        raise HTTPException(400, "File is empty")

    # Parse records
    records: list[OMRValidationRecord] = []
    duplicate_ids: set[str] = set()
    seen_ids: set[str] = set()
    errors: list[str] = []
    invalid_student_ids: list[str] = []

    import re as _re

    for line_num, line in enumerate(lines, 1):
        parts = [p.strip() for p in line.split(",")]
        if len(parts) < 2:
            errors.append(f"Row {line_num}: insufficient columns (expected at least 2)")
            continue

        marker = parts[0]
        omr_id = parts[1]

        if marker != "x":
            errors.append(f"Row {line_num}: invalid marker '{marker}' (expected 'x')")
            continue

        if not omr_id:
            errors.append(f"Row {line_num}: empty OMR ID")
            continue

        # Check for duplicates
        if omr_id in seen_ids:
            duplicate_ids.add(omr_id)
            errors.append(f"Row {line_num}: duplicate OMR ID '{omr_id}'")
            continue

        seen_ids.add(omr_id)

        # Parse answers using question type at each position
        answers_raw = [p for p in parts[2:] if p != ""]
        try:
            answers: list[int | str] = []
            for i, ans in enumerate(answers_raw):
                qt = qtypes[i] if i < len(qtypes) else ""
                if ans == "-1000000":
                    answers.append(-1000000)
                elif qt == "MCQ":
                    # MCQ: concatenated letters "BCD" or digits "234" → pipe-separated "2|3|4"
                    answers.append(_normalise_answer_key(ans, "MCQ"))
                elif qt == "DECIMAL" or "." in ans:
                    float(ans)  # validate; raises ValueError if not a number
                    answers.append(ans)
                else:
                    # SCQ / INT: single letter or digit
                    converted = _normalise_answer_key(ans, "SCQ")
                    answers.append(int(converted))
        except ValueError:
            errors.append(f"Row {line_num} ({omr_id}): invalid answer value '{ans}'")
            invalid_student_ids.append(omr_id)
            continue

        records.append(OMRValidationRecord(omr_id=omr_id, answers=answers))

    # Validate students exist and are enrolled in the specified branch
    if records:
        submitted_omr_ids = {r.omr_id for r in records}
        enrolled_students = get_enrolled_students_for_exam(db, exam_id, branch_id=branch_id)
        enrolled_map: dict[str, Student] = {}
        conflicting_omr_ids: set[str] = set()
        for student in enrolled_students:
            omr_id = student.admission_no[-7:]
            if omr_id in conflicting_omr_ids:
                continue
            if omr_id in enrolled_map:
                conflicting_omr_ids.add(omr_id)
                enrolled_map.pop(omr_id, None)
            else:
                enrolled_map[omr_id] = student

        for omr_id in sorted(conflicting_omr_ids):
            invalid_student_ids.append(omr_id)
            errors.append(f"OMR ID '{omr_id}' matches multiple enrolled students; admission numbers must have unique last 7 characters")

        for record in records:
            if record.omr_id not in enrolled_map:
                invalid_student_ids.append(record.omr_id)
                errors.append(f"OMR ID '{record.omr_id}' not enrolled in {exam.program.name} / {exam.class_.name}")

        # Find missing students
        missing = set(enrolled_map.keys()) - submitted_omr_ids
        missing_list = sorted(list(missing))
        absent_students = [
            OMRAbsentStudent(
                omr_id=omr_id,
                admission_no=enrolled_map[omr_id].admission_no,
                name=enrolled_map[omr_id].name,
            )
            for omr_id in missing_list
        ]
    else:
        missing_list = []
        absent_students = []

    return OMRValidationSummary(
        valid_count=len(records) - len(set(invalid_student_ids)),
        duplicate_ids=sorted(list(duplicate_ids)),
        invalid_student_ids=sorted(list(set(invalid_student_ids))),
        missing_students=missing_list,
        absent_students=absent_students,
        errors=errors,
        file_records=records,
        program_id=exam.program_id,
        class_id=exam.class_id,
    )


@router.post("/{exam_id}/results/save", dependencies=[admin_only])
def save_omr_results(exam_id: int, data: OMRUploadConfirm, db: DbSession):
    """Save validated OMR results to database and record upload summary."""
    from app.models.exam_upload_log import ExamUploadLog
    from datetime import datetime, timezone

    exam = db.get(Exam, exam_id)
    if not exam:
        raise HTTPException(404, "Exam not found")
    if exam.status != EXAM_STATUS_PUBLISHED:
        raise HTTPException(409, "OMR results can only be uploaded to published exams")

    if exam_id != data.exam_id:
        raise HTTPException(400, "Exam ID mismatch")

    enrolled_students = get_enrolled_students_for_exam(db, exam_id, branch_id=data.branch_id)
    enrolled_map: dict[str, Student] = {}
    conflicting_omr_ids: set[str] = set()
    for student in enrolled_students:
        omr_id = student.admission_no[-7:]
        if omr_id in conflicting_omr_ids:
            continue
        if omr_id in enrolled_map:
            conflicting_omr_ids.add(omr_id)
            enrolled_map.pop(omr_id, None)
        else:
            enrolled_map[omr_id] = student

    # ── Save valid records, skip invalid ones ─────────────────────────────
    errors: list[str] = []
    saved = 0

    for record in data.records:
        if record.omr_id in conflicting_omr_ids:
            errors.append(f"OMR ID '{record.omr_id}' matches multiple enrolled students")
            continue
        if record.omr_id not in enrolled_map:
            errors.append(f"OMR ID '{record.omr_id}' not found in enrolled list")
            continue
        student = enrolled_map[record.omr_id]
        answers_str = ",".join(str(a) for a in record.answers)
        try:
            upsert_exam_result(db, exam_id, student.id, answers_str)
            saved += 1
        except Exception as exc:
            errors.append(f"Failed to save '{record.omr_id}': {exc}")

    # Upsert upload log (per exam + branch)
    log = db.scalar(
        select(ExamUploadLog).where(
            ExamUploadLog.exam_id == exam_id,
            ExamUploadLog.branch_id == data.branch_id,
        )
    )
    if log:
        log.uploaded_at = datetime.now(timezone.utc)
        log.valid_count = data.valid_count
        log.absent_count = data.absent_count
        log.duplicate_count = data.duplicate_count
        log.invalid_count = data.invalid_count
        log.absent_list = ",".join(data.absent_list)
        log.file_name = data.file_name
    else:
        log = ExamUploadLog(
            exam_id=exam_id,
            branch_id=data.branch_id,
            uploaded_at=datetime.now(timezone.utc),
            valid_count=data.valid_count,
            absent_count=data.absent_count,
            duplicate_count=data.duplicate_count,
            invalid_count=data.invalid_count,
            absent_list=",".join(data.absent_list),
            file_name=data.file_name,
        )
        db.add(log)

    db.commit()

    return {
        "saved": saved,
        "errors": errors,
    }
